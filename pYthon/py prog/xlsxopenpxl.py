# -*- coding: utf-8 -*-
"""
Created on Fri Aug  9 19:08:34 2019

@author: faculty
"""

import openpyxl as oxl
file_name = "D:\\April\\realEstate_trans.xlsx"
workbook = oxl.load_workbook(file_name)


print(workbook.sheetnames)
sheet = workbook.active
#i=0
lable = []
for row in sheet.iter_rows():
    #i+=1
    #if i>1:
        #break
    for cell in row:
        lable.append(cell.value)
print(lable)
